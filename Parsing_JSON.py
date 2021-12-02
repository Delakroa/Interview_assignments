import openpyxl
import json

json_data = json.load(open('test1.json', encoding='utf-8'))
json_data2 = json.load(open('test2.json', encoding='utf-8'))
json_data3 = json.load(open('test3.json', encoding='utf-8'))

row = []
for header in json_data['headers']:
    text = header['properties']
    QuickInfo = text['QuickInfo']
    row.append(QuickInfo)
    # print(type(row))

# Создаём рабочую книгу
book = openpyxl.Workbook()
# По умолчанию создаётся с таблицей Sheet
# print(book.sheetnames)
book.remove(book.active)

# Позиционирование на конкретном листе
# в который мы осуществляем запись
sheet_1 = book.create_sheet('Лист1')  # по умолчанию это первый лист
sheet_2 = book.create_sheet('Лист2')
sheet_3 = book.create_sheet('Лист3')

# Обращение к координатам ячейки
sheet_1['A1'] = str(row[0:1])
sheet_1['B1'] = str(row[1:2])
sheet_1['C1'] = str(row[2:3])
sheet_1['D1'] = str(row[3:4])
sheet_2['A1'] = str(row[0:1])
sheet_2['B1'] = str(row[1:2])
sheet_2['C1'] = str(row[2:3])
sheet_2['D1'] = str(row[3:4])
sheet_3['A1'] = str(row[0:1])
sheet_3['B1'] = str(row[1:2])
sheet_3['C1'] = str(row[2:3])
sheet_3['D1'] = str(row[3:4])


data_sample = []
data_sample2 = []
data_sample3 = []
rows = 2
for header in json_data['values']:
    text = header['properties']
    data_samples = text['Text']
    data_sample.append(data_samples)
print(data_sample)

for header2 in json_data2['values']:
    text2 = header2['properties']
    data_samples2 = text2['Text']
    data_sample2.append(data_samples2)
print(data_sample2)

for header3 in json_data3['values']:
    text3 = header3['properties']
    data_samples3 = text3['Text']
    data_sample3.append(data_samples3)
    # print(data_sample3)
    # Обращение по координатам [ряд][колонка]
    sheet_1[rows][0].value = str(data_sample[0:1])
    sheet_1[rows][1].value = str(data_sample[2:3])
    sheet_1[rows][2].value = str(data_sample[4:5])
    sheet_1[rows][3].value = str(data_sample[6:7])
    sheet_2[rows][0].value = str(data_sample2[0:1])
    sheet_2[rows][1].value = str(data_sample2[2:3])
    sheet_2[rows][2].value = str(data_sample2[4:5])
    sheet_2[rows][3].value = str(data_sample2[6:7])
    sheet_3[rows][0].value = str(data_sample3[0:1])
    sheet_3[rows][1].value = str(data_sample3[2:3])
    sheet_3[rows][2].value = str(data_sample3[4:5])
    sheet_3[rows][3].value = str(data_sample3[6:7])
    rows += 1

# Чтобы изменения вступили в силу
# сохраняем книгу
book.save("my_book.xlsx")
book.close()
